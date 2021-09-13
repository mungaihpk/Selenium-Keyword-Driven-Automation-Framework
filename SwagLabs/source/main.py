from methods import get_web_element, open_browser, input_data, verify_text_present, verify_element_present, verify_text_contains, verify_text, add_to_shopping_cart, remove_from_cart
from pathlib import Path
from openpyxl import load_workbook
from selenium.common.exceptions import ElementNotInteractableException
import time
import calendar
import datetime


def read_steps():
    # set file path
    path = Path(__file__).parent / "../data_engine/keywords.xlsx"

    # load excel file
    workbook = load_workbook(path)

    # activate demo.xlsx
    sheet = workbook["Steps"]

    rows = sheet.max_row

    test_steps = []
    count = 1

    for row in sheet.iter_rows(min_row=2):
        step = row[0].value
        keyword = row[2].value
        page_object = row[3].value
        identifier = row[4].value
        value = row[5].value
        input_value = row[6].value

        if step is not None:
            count += 1
            test_step = {"step": step, "keyword": keyword, "object": page_object,
                         "identifier": identifier, "value": value, "input": input_value, "row": count}
            print(count)
            test_steps.append(test_step)
        else:
            break

    for item in test_steps:
        print(item["step"])

    return {"test_steps": test_steps, "sheet": sheet, "workbook": workbook}


def execute_test(test_step):
    pass


if __name__ == '__main__':
    browser = None
    read_excel = read_steps()
    test_steps = read_excel["test_steps"]
    sheet = read_excel["sheet"]
    workbook = read_excel["workbook"]

    for test_step in test_steps:

        step = test_step["step"]
        keyword = test_step["keyword"]
        page_object = test_step["object"]
        identifier = test_step["identifier"]
        value = test_step["value"]
        input_value = test_step["input"]
        row = test_step["row"]

        if keyword == "OpenBrowser":
            browser = open_browser(value)
            sheet["H"+str(row)] = "Passed"

        elif keyword == "MaximizeWindow":
            browser.maximize_window()
            sheet["H"+str(row)] = "Passed"
        elif keyword == "MinimizeWindow":
            browser.minimize_window()
            sheet["H"+str(row)] = "Passed"
        elif keyword == "NavigateTo":
            browser.get(input_value)
            sheet["H"+str(row)] = "Passed"
        elif keyword == "InputData":
            element = get_web_element(browser, identifier, value)
            input_data(browser, element, input_value)
            sheet["H"+str(row)] = "Passed"
        elif keyword == "Click":
            element = get_web_element(browser, identifier, value)
            try:
                element.click()
                sheet["H"+str(row)] = "Passed"
            except ElementNotInteractableException:
                print("The element is present but hidden")
                sheet["H"+str(row)] = "Failed"
        elif keyword == "VerifyTextPresent":
            pass
        elif keyword == "AddToShoppingCart":
            if str(input_value).strip() is not "":
                add_to_cart = add_to_shopping_cart(browser, input_value)
                if add_to_cart["added"]:
                    sheet["H"+str(row)] = "Passed"
                    sheet["I"+str(row)] = ""
                else:
                    sheet["H"+str(row)] = "Failed"
                    sheet["I"+str(row)] = add_to_cart["message"]
            else:

                add_to_cart = add_to_shopping_cart(browser)
                if add_to_cart["added"]:
                    sheet["H"+str(row)] = "Passed"
                    sheet["I"+str(row)] = ""
                else:
                    sheet["H"+str(row)] = "Failed"
                    sheet["I"+str(row)] = add_to_cart["message"]
        elif keyword == "Delay":
            time.sleep(input_value)
            sheet["H"+str(row)] = "Passed"
        elif keyword == "Back":
            browser.back()
        elif keyword == "VerifyElementPresent":
            is_element_present = verify_element_present(
                browser, identifier, value)
            if is_element_present:
                sheet["H"+str(row)] = "Passed"
            else:
                # fail the test
                sheet["H"+str(row)] = "Failed"
                sheet["I"+str(row)] = add_to_cart["message"]
        elif keyword == "RemoveFromShoppingCart":
            if str(input_value).strip() is not "":
                remove_from_cart = remove_from_cart(browser, input_value)
                if remove_from_cart["removed"]:
                    sheet["H"+str(row)] = "Passed"
                    sheet["I"+str(row)] = ""
                else:
                    sheet["H"+str(row)] = "Failed"
                    sheet["I"+str(row)] = remove_from_cart["message"]
            else:
                sheet["H"+str(row)] = "Failed"
                sheet["I"+str(row)] = "You need to specify the name of the product to remove"

    workbook.save(Path(__file__).parent /
                  '../results/results_{}.xlsx'.format(
                      calendar.timegm(time.gmtime())))

    # workbook.save(root_path)
    # browser.close()
